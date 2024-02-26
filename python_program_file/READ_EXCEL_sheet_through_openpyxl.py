import openpyxl

wb=openpyxl.load_workbook("F:\openpyxl.xlsx")
sh1=wb['malaya']
row=sh1.max_row
column=sh1.max_column
#print(row)
#print(column)

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value)
# add data
sh1.cell(row=6,column=1,value="tusar")
sh1.cell(row=6,column=2,value="balasore")
sh1.cell(row=6,column=3,value="7")

sh1.cell(row=7,column=1,value="sanapua")
sh1.cell(row=7,column=2,value="chanchina")
sh1.cell(row=7,column=3,value="102")

wb.save('report.xlsx')
