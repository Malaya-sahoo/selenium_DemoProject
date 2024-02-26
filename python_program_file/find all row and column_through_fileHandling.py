import openpyxl

wb=openpyxl.load_workbook('F:\collegefile.xlsx')

#wk=wb.sheetnames
#print(wk)
#sh=wb.active.title
#print(sh)
sh=wb['Name']
row=sh.max_row
columns=sh.max_column
#print("total rows are)-" + str(row))
#print("total columns are)-" +

for r in sh['A1':'C4']:
    for c in r:
        print(c.value)

#for i in range(1,row+1):
#    for j in range(1,columns+1):
#        c=sh.cell(i,j)
#        print(c.value)